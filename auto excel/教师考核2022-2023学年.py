import xlrd
import openpyxl

wb = openpyxl.Workbook
data = xlrd.open_workbook("1.xlsx")
sheet1 = data.sheet_by_index(0)
sheet2 = data.sheet_by_index(1)

wb = openpyxl.Workbook()
ws = wb.worksheets[0]

for i in range(sheet1.nrows):
    for j in range(sheet2.nrows):
        if (sheet1.cell(i, 0).value == sheet2.cell(j, 0).value) and (sheet1.cell(i, 1).value == sheet2.cell(j, 1).value):
            ws.cell(row=i+1, column=1, value=sheet1.cell(i, 0).value)
            ws.cell(row=i+1, column=2, value=sheet1.cell(i, 1).value)
            ws.cell(row=i+1, column=3, value=sheet1.cell(i, 2).value)
            ws.cell(row=i+1, column=4, value=sheet2.cell(j, 2).value)
        else:
            ws.cell(row=i+1, column=1, value=sheet1.cell(i, 0).value)
            ws.cell(row=i+1, column=2, value=sheet1.cell(i, 1).value)
            ws.cell(row=i+1, column=3, value=sheet1.cell(i, 2).value)
wb.save('2.xlsx')
wb.close()
