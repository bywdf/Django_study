import openpyxl

wb = openpyxl.load_workbook('原始.xlsx')
ws = wb.worksheets[0]

wb2 = openpyxl.Workbook()
ws2 = wb2.active

# for row in ws.iter_rows(min_row=1,max_row=11):
#     for cell in row:
#         print(cell.value, end="")
#     print()
a_list = ['考号','姓名','班级','语文','数学','外语','选课1','成绩1','选课2','成绩2','选课3','成绩3','总成绩','总位次','选课1位次','选课2位次','选课三位次']

ws2.append(a_list)

for i in range(2,12):
    rowlist = []
    for j in range(1,18):
        a = ws.cell(i,j).value
        if 7 <= j <= 12: 
            if a != None:
                rowlist.append(ws.cell(1,j).value)
                rowlist.append(a)
        else:
            rowlist.append(a)
    ws2.append(rowlist)
    

wb2.save('1.xlsx')