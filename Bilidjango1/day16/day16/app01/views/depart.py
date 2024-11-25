from django.shortcuts import render, redirect, HttpResponse
from app01 import models

from app01.utils.pagination import Pagination

# Create your views here.

def depart_list(request):
    '''部门列表'''
    queryset = models.Department.objects.all()
    return render(request, 'depart_list.html', {'queryset': queryset})


def depart_add(request):
    '''添加部门'''
    if request.method == 'GET':
        return render(request, 'depart_add.html')

    title = request.POST.get('title')
    models.Department.objects.create(title=title)
    return redirect('/depart/list/')


def depart_delete(request):
    '''删除部门'''
    nid = request.GET.get('nid')
    models.Department.objects.filter(id=nid).delete()
    return redirect('/depart/list/')


def depart_edit(request, nid):
    '''修改部门'''
    # 通过nid获取数据，获取到的是一个列表对象，获取第一个
    if request.method == 'GET':
        row_object = models.Department.objects.filter(id=nid).first()
        return render(request, 'depart_edit.html',{'row_object': row_object})
    title = request.POST.get('title')
    models.Department.objects.filter(id=nid).update(title=title)
    return redirect('/depart/list')


def depart_multi(request):
    '''批量上传 (excel)'''
    from openpyxl import load_workbook
    # django.core.files.uploadedfile.InMemoryUploadedFile
    
    # 1.获取用户上传的对象文件
    file_object = request.FILES.get("exc")
    # print(type(file_object))
    
    # 2.对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    
    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
    
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)
    
    # with open(file_object.name, mode='wb') as f:
    #     for chunk in file_object:
    #         f.write(chunk)
    
    return redirect("/depart/list/")