from django.shortcuts import render, redirect
from cjcx import models
import openpyxl

from cjcx.utils.pagination import Pagination

# Create your views here.


def list(request):
    '''所有学生'''

    data_dict = {}
    search_data = request.GET.get('q', '')
    if search_data:
        data_dict['zhunkaozheng__contains'] = search_data
    # models.PrettyNum.objects.filter(**data_dict)

    queryset = models.Chengji.objects.filter(**data_dict)

    page_object = Pagination(request, queryset)

    context = {
        'search_data': search_data,
        'queryset': page_object.page_queryset,     # 分页数据
        'page_string': page_object.html()         # 页码
    }

    return render(request, 'list.html', context)


def jieguo(request, nid):
    '''成绩查询'''
    queryset = models.Chengji.objects.filter(id=nid)
    return render(request, 'jieguo.html', {'queryset': queryset})


def add(request):
    '''添加学生'''

    wb = openpyxl.load_workbook(r'D:\myGithub\Django_study\gkcjcx\cjcx\1.xlsx')
    ws = wb.worksheets[0]

    for i in range(2, 12):
        models.Chengji.objects.create(
            zhunkaozheng=ws.cell(i, 1).value,
            xingming=ws.cell(i, 2).value,
            banji=ws.cell(i, 3).value,

            yuwen_score=ws.cell(i, 4).value,
            shuxue_score=ws.cell(i, 5).value,
            waiyu_score=ws.cell(i, 6).value,

            xuankao_1=ws.cell(i, 7).value,
            xuankao_1_score=ws.cell(i, 8).value,
            xuankao_1_weici=ws.cell(i, 15).value,
            xuankao_2=ws.cell(i, 9).value,
            xuankao_2_score=ws.cell(i, 10).value,
            xuankao_2_weici=ws.cell(i, 16).value,
            xuankao_3=ws.cell(i, 11).value,
            xuankao_3_score=ws.cell(i, 12).value,
            xuankao_3_weici=ws.cell(i, 17).value,

            total_score=ws.cell(i, 13).value,
            total_weici=ws.cell(i, 14).value,
        )

    return redirect('/list/')


def delete(request):
    '''清除所有数据'''

    models.Chengji.objects.all().delete()
    return redirect('/list/')
